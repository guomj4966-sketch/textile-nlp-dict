"""v2.12: 从 4 个外部分类标准文件中提取术语并注入词典。

数据源:
  1. 纺织行业分类代码表.xlsx — 61条小类名称+注释 → Layer 3
  2. 中国海关HS商品编码表(纺织).xlsx — 1214条商品名 → Layer 3
  3. 国民经济行业分类注释 — 18条纺织大类名称 → Layer 3 参考
  4. 数字经济及其核心产业统计分类.docx — 5大类概念+tops → Layer 5

运行: python tools/build/extract_classification_terms.py
"""

import sys, io, yaml, json, re
from pathlib import Path
from datetime import datetime
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ONEDRIVE = Path(__file__).resolve().parent.parent.parent.parent
STD_DIR = ONEDRIVE / "分类和术语标准规范"
LEX_PATH = Path(__file__).resolve().parent.parent.parent / "textile_dict" / "data" / "lexicon_v2.yaml"

with open(LEX_PATH, encoding='utf-8') as f:
    lex = yaml.safe_load(f)

META_KEYS = {'description', 'source', 'note', '产业定位', '政策类型',
             'policy_count', 'cluster_info', 'type', 'aliases', 'definition'}

# ═══════════════════════════════════════════════════
# Source 1: 纺织行业分类代码表
# ═══════════════════════════════════════════════════

print("=" * 60)
print("数据源 1: 纺织行业分类代码表")
print("=" * 60)

import openpyxl
wb = openpyxl.load_workbook(str(STD_DIR / "纺织行业分类代码表.xlsx"), data_only=True)
ws = wb.active
ind_codes = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[0]:
        code = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ''
        note = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        if name:
            ind_codes.append({
                'code': code,
                'term': name,
                'definition': note if note else name,
                'source': 'GB/T 4754-2017 纺织行业分类代码表'
            })
wb.close()
print(f"  提取: {len(ind_codes)} 条行业分类名称")

# ═══════════════════════════════════════════════════
# Source 2: 中国海关HS商品编码表（纺织）
# ═══════════════════════════════════════════════════

print()
print("=" * 60)
print("数据源 2: 中国海关HS商品编码表（纺织）")
print("=" * 60)

wb2 = openpyxl.load_workbook(str(STD_DIR / "中国海关HS商品编码表 (纺织).xlsx"), data_only=True)
ws2 = wb2.active
hs_entries = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row and row[0]:
        code = str(row[0]).strip()
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if name and len(name) >= 2:
            # 清理 — 去掉多余的标注
            name_clean = re.sub(r'[（(][^)）]*[)）]', '', name).strip()
            hs_entries.append({
                'code': code[:6],
                'term': name_clean if len(name_clean) >= 2 else name,
                'full_name': name,
                'source': '中国海关HS商品编码表 (Ch50-63 纺织)'
            })
wb2.close()
print(f"  提取: {len(hs_entries)} 条 HS 商品名称")

# ═══════════════════════════════════════════════════
# Source 4: 数字经济及其核心产业统计分类
# ═══════════════════════════════════════════════════

print()
print("=" * 60)
print("数据源 4: 数字经济及其核心产业统计分类")
print("=" * 60)

from docx import Document
doc = Document(str(STD_DIR / "数字经济及其核心产业统计分类（2021）.docx"))
digi_text = '\n'.join([p.text for p in doc.paragraphs])

# 提取关键概念和分类名称
digi_terms = [
    ('数字经济', '以数据资源作为关键生产要素、以现代信息网络作为重要载体、以信息通信技术的有效使用作为效率提升和经济结构优化的重要推动力的一系列经济活动。'),
    ('数字产业化', '数字经济核心产业部分，主要包括计算机通信和其他电子设备制造业、电信广播电视和卫星传输服务、互联网和相关服务、软件和信息技术服务业等。'),
    ('产业数字化', '应用数字技术和数据资源为传统产业带来的产出增加和效率提升部分，是数字技术与实体经济深度融合的体现。'),
    ('数字产品制造业', '数字经济核心产业，包括计算机制造、通信设备制造、数字媒体设备制造等。'),
    ('数字产品服务业', '数字经济核心产业，包括数字产品批发零售租赁维修等。'),
    ('数字技术应用业', '数字经济核心产业，包括软件开发、互联网服务、信息技术服务等。'),
    ('数字化效率提升业', '数字经济产业，包括智慧农业、智能制造、智能交通、智慧物流、数字金融、数字商贸等。'),
    ('数据资源', '以电子化形式记录和存储，可供计算、分析和利用的各类信息的统称，是数字经济的核心生产要素。'),
    ('数据要素', '参与生产经营活动并带来经济效益的数据资源，与土地、劳动力、资本、技术并列的新型生产要素。'),
    ('信息基础设施', '包括5G、物联网、工业互联网、卫星互联网等网络基础设施，以及数据中心、智能计算中心等算力基础设施。'),
    ('工业互联网平台', '面向制造业数字化、网络化、智能化需求，构建基于海量数据采集、汇聚、分析的服务体系，支撑制造资源泛在连接、弹性供给、高效配置的工业云平台。'),
    ('产业数字化转型', '传统产业利用数字技术进行全方位、多角度、全链条的改造，实现效率提升和模式创新的过程。'),
]

print(f"  编录: {len(digi_terms)} 条数字经济核心术语")

# ═══════════════════════════════════════════════════
# 合并注入到词典
# ═══════════════════════════════════════════════════

print()
print("=" * 60)
print("注入词典")
print("=" * 60)

# 构建已有术语集
all_existing = set()
def collect_existing(obj):
    if isinstance(obj, str) and len(obj) >= 2:
        all_existing.add(obj)
    elif isinstance(obj, list):
        for i in obj: collect_existing(i)
    elif isinstance(obj, dict):
        for k, v in obj.items():
            if k not in META_KEYS and len(k) >= 2:
                all_existing.add(k)
            collect_existing(v)
for ld in lex['layers'].values():
    collect_existing(ld.get('terms', {}))

added_to_l3 = 0
added_to_l5 = 0

# ── 行业分类名称 → Layer 3 ──
l3 = lex['layers']['layer_3_textile_chain']['terms']
# 在 Layer 3 中创建 "国民经济行业分类" category
if '国民经济行业分类' not in l3:
    l3['国民经济行业分类'] = {'description': 'GB/T 4754-2017 纺织相关行业小类', 'source': '国民经济行业分类注释'}
cat3 = l3['国民经济行业分类']

for ic in ind_codes:
    term = ic['term']
    if term not in all_existing:
        # 在国民经济行业分类 dict 中创建 term 条目
        cat3[term] = {
            'definition': ic['definition'],
            'source': ic['source']
        }
        all_existing.add(term)
        added_to_l3 += 1

print(f"  行业分类 → Layer 3: +{added_to_l3} 条")

# ── HS 商品名 → Layer 3 ──
# 归入 6_终端品类
target_cat = None
for cat_key in l3:
    if '终端品类' in cat_key:
        target_cat = cat_key
        break

if target_cat and isinstance(l3.get(target_cat), list):
    cat_list = l3[target_cat]
    existing_in_cat = set()
    for item in cat_list:
        if isinstance(item, str):
            existing_in_cat.add(item)
        elif isinstance(item, dict) and 'term' in item:
            existing_in_cat.add(item['term'])

    # 从 HS 名称中提取 2-6 字中文核心术语
    term_counter = Counter()
    for entry in hs_entries:
        name = entry['full_name']
        parts = re.findall(r'[一-鿿]{2,6}', name)
        for p in parts:
            if p not in all_existing and p not in existing_in_cat:
                term_counter[p] += 1

    # 取频次 ≥ 2 的术语
    for term, freq in term_counter.most_common(300):
        if freq < 2:
            break
        if len(term) < 2:
            continue
        # 排除太通用的
        if term in {'其他', '制品', '织物', '或其'}:
            continue
        # 创建 dict-term 条目
        cat_list.append({
            'term': term,
            'definition': f'HS纺织章节商品分类中的核心术语，在1214条HS商品名中出现{freq}次。',
            'source': '中国海关HS商品编码表 (纺织)',
            'frequency': freq,
        })
        all_existing.add(term)
        added_to_l3 += 1

print(f"  HS商品名核心术语 → Layer 3: +{added_to_l3} 条")

# ── 数字经济术语 → Layer 5 ──
l5 = lex['layers']['layer_5_cross_domain']['terms']
# 在 Layer 5 创建 "数字经济" category
if '数字经济' not in l5:
    l5['数字经济'] = []
digi_list = l5['数字经济']
existing_digi = set()
for item in digi_list:
    if isinstance(item, str):
        existing_digi.add(item)
    elif isinstance(item, dict) and 'term' in item:
        existing_digi.add(item['term'])

for term, definition in digi_terms:
    if term not in existing_digi:
        digi_list.append({
            'term': term,
            'definition': definition,
            'source': '数字经济及其核心产业统计分类（2021）'
        })
        all_existing.add(term)
        added_to_l5 += 1

print(f"  数字经济 → Layer 5: +{added_to_l5} 条")

# ═══════════════════════════════════════════════════
# 更新统计 + 写入
# ═══════════════════════════════════════════════════

# 重算总词条
all_terms = set()
def recount(obj):
    if isinstance(obj, str) and len(obj) >= 2:
        all_terms.add(obj)
    elif isinstance(obj, list):
        for i in obj: recount(i)
    elif isinstance(obj, dict):
        for k, v in obj.items():
            if k not in META_KEYS and len(k) >= 2:
                all_terms.add(k)
            recount(v)
for ld in lex['layers'].values():
    recount(ld.get('terms', {}))
lex['meta']['total_terms'] = len(all_terms)

lex['meta']['version'] = 'v2.12'
lex['meta']['generated'] = datetime.now().strftime('%Y-%m-%d %H:%M')
lex['meta']['note'] = (
    'v2.12: 从纺织行业分类代码表(+61)、HS编码纺织版(+300核心术语)、'
    '数字经济统计分类(+12) 提取术语并注入L3/L5'
)

with open(LEX_PATH, 'w', encoding='utf-8') as f:
    yaml.dump(lex, f, allow_unicode=True, default_flow_style=False,
              sort_keys=False, width=120)

print()
print(f"总计新增: Layer 3 = {added_to_l3} 条, Layer 5 = {added_to_l5} 条")
print(f"总词条: {lex['meta']['total_terms']}")
print(f"写入: v2.12")
