"""从外部资料源批量提取纺织领域候选术语，去重并分类。

数据源:
  1. HS 商品编码表 (Ch50-63 纺织品类) — 819 条产品名称
  2. 前五批专精特新特色产品 — 397 条产品名称
  3. GB/T 标准文件名 — 15 份纺织相关标准
  4. 国民经济行业分类（纺织相关大类） — 行业分类名称

运行:
    PYTHONIOENCODING=utf-8 python scripts/nlp_dict/enrich_from_sources.py

输出:
    scripts/nlp_dict/data/source_enrichment.csv  — 候选术语 + 来源 + 建议分类
"""

import sys
import re
import csv
import io
import argparse
from pathlib import Path
from collections import Counter

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import yaml

# ============================================================
# 配置
# ============================================================

DATA_DIR = Path(__file__).parent / "data"
LEXICON_PATH = DATA_DIR / "lexicon_v2.yaml"
OUTPUT_PATH = DATA_DIR / "source_enrichment.csv"

# 从脚本位置向上4级到 OneDrive 根目录，再进入各自文件夹
_ONEDRIVE = Path(__file__).parent.parent.parent.parent
HS_CODE_PATH = _ONEDRIVE / "分类和术语标准规范" / "中国海关HS商品编码表.csv"
GB50514_PATH = _ONEDRIVE / "分类和术语标准规范" / "GB 50514-2009 非织造布工厂设计规范.pdf"
PRODUCTS_PATH = _ONEDRIVE / "专精特新平台" / "前五批纺织行业专精特新特色产品.xlsx"
PATENTS_PATH = _ONEDRIVE / "专精特新平台" / "专精特新企业专利信息2025.xlsx"

# HS Code 纺织章节范围
TEXTILE_HS_CHAPTERS = [str(i) for i in range(50, 64)]  # Ch50-Ch63


# ============================================================
# 数据源 1: HS 商品编码表
# ============================================================

def extract_hs_terms():
    """从 HS Code Ch50-63 提取纺织品品类名称。

    策略: 取完整的 10 位码商品名称，同时分解出其中含有的 2-4 字核心术语。
    """
    entries = []  # [(code, full_name)]
    with open(HS_CODE_PATH, encoding='gbk') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            code = row[0].strip()
            if len(code) >= 2 and code[:2] in TEXTILE_HS_CHAPTERS:
                name = row[1].strip()
                # 基本清洗
                name = re.sub(r'[，、；；].*', '', name)  # 截断到第一个标点
                name = re.sub(r'\(.*?\)', '', name)       # 去括号备注
                name = re.sub(r'\d+', '', name)            # 去数字
                name = name.strip()
                if len(name) >= 3:
                    entries.append((code[:2], name))

    # 按二级章分组
    by_chapter = {}
    for ch, name in entries:
        by_chapter.setdefault(ch, []).append(name)

    print(f"   HS Code Ch50-63: {len(entries)} 条商品名称")
    for ch in sorted(by_chapter):
        print(f"     Ch{ch}: {len(by_chapter[ch])} 条")

    # 从完整名称中分离核心术语 (2-6 字的中文片段)
    core_terms = Counter()
    for _, name in entries:
        # 分割长名: 以 "及""或""和""、" 为分割
        parts = re.split(r'[及或和、．，,]', name)
        for part in parts:
            part = part.strip()
            # 去掉前缀修饰 (如 "其他""未列名""非")
            part = re.sub(r'^(其他|未列名|未列明|不论是否|不适于|非零售|非供零售用)', '', part)
            part = re.sub(r'(用|制|的)$', '', part)
            part = part.strip()
            if 2 <= len(part) <= 8 and re.match(r'^[一-鿿]{2,8}$', part):
                core_terms[part] += 1

    # 取频次 >= 1 的核心术语（HS 名称中的术语重复出现说明是通用概念）
    hs_terms = [t for t, f in core_terms.items() if f >= 1]
    print(f"     核心术语 (2-8字): {len(hs_terms)} 个")

    return entries, hs_terms


# ============================================================
# 数据源 2: 专精特新特色产品
# ============================================================

def extract_product_terms():
    """从前五批特色产品 Excel 提取产品名称中的新术语。"""
    try:
        import openpyxl
    except ImportError:
        print("   ⚠️ openpyxl 未安装，跳过产品术语提取")
        return [], []

    wb = openpyxl.load_workbook(PRODUCTS_PATH, data_only=True)
    ws = wb['前5批特色产品']

    products = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue  # skip header
        # B列: 特色产品名称, D列: 其他主打产品
        for col_idx in [1, 3]:
            if len(row) > col_idx and row[col_idx]:
                name = str(row[col_idx]).replace('_x000D_', '').replace('\n', ' ').strip()
                if name and len(name) >= 3:
                    products.append(name)

    print(f"   专精特新特色产品: {len(products)} 条产品名称")

    # 提取 2-6 字中文术语 (去除常见修饰词)
    core_terms = Counter()
    for name in products:
        chinese_parts = re.findall(r'[一-鿿]{2,8}', name)
        for part in chinese_parts:
            # 跳过纯数字/单位/品牌名
            part = part.strip()
            if re.match(r'^[零一二三四五六七八九十百千万亿]+$', part):
                continue
            if len(part) >= 2:
                core_terms[part] += 1

    product_terms = [t for t, f in core_terms.most_common(300) if f >= 1]
    print(f"     核心术语 (2-8字): {len(product_terms)} 个 (取 Top 300)")

    return products, product_terms


# ============================================================
# 数据源 3: 行业分类标准中的纺织段落
# ============================================================

def extract_industry_class_terms():
    """从国民经济行业分类注释中提取纺织相关分类名。"""
    try:
        import openpyxl
    except ImportError:
        print("   ⚠️ openpyxl 未安装，跳过行业分类提取")
        return []

    # 纺织大类编码范围
    textile_ranges = [
        ('17', '纺织业'),
        ('18', '纺织服装、服饰业'),
        ('28', '化学纤维制造业'),
    ]

    terms = []
    try:
        wb = openpyxl.load_workbook(
            _ONEDRIVE / "分类和术语标准规范" / "《2017国民经济行业分类注释》（按第1号修改单修订）.xlsx",
            data_only=True
        )
        ws = wb['Sheet1']

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or not row[0]:
                continue
            code = str(row[0]).strip()
            # Check if code falls under textile segments
            for prefix, label in textile_ranges:
                if code.startswith(prefix):
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    if name and len(name) >= 3:
                        terms.append((code, name, label))
                    break

        wb.close()
    except Exception as e:
        print(f"   ⚠️ 行业分类读取异常: {e}")

    print(f"   国民经济行业分类: {len(terms)} 条纺织相关")
    class_names = [t[1] for t in terms]
    return class_names


# ============================================================
# 数据源 4: GB/T 标准名称中隐含的术语体系
# ============================================================

def extract_standard_terms():
    """从 GB/T 标准的主题领域提取分类术语。"""
    # 标准文件名编码的知识
    standard_domains = [
        # 丝绸
        '桑蚕茧', '秋茧', '丝绵', '桑蚕丝弹力织物', '桑波缎', '顺纡绉',
        '苏绣', '蜀绣', '湘绣', '粤绣', '丝绸服装', '丝绸家居服', '丝绸旗袍',
        '丝针织服装', '丝绸围巾', '丝绸领带', '蚕丝绒毯', '丝绸床上用品',
        '蚕丝凉席', '丝绸眼罩', '丝绸包', '丝绸书', '丝绸画', '丝绸册页',
        '丝绒织物', '漳绒', '漳缎', '宋锦', '蜀锦', '壮锦', '缂丝', '莨纱绸', '杭罗',
        '原料精练', '圆网印花', '扎染', '蜡染', '植物染料染色', '柔软整理', '预缩整理',
        '织物后整理',
        # 缝制机械
        '缝制机械', '功能部件', '工业缝纫机', '刺绣机', '绗绣机', '铺布裁剪设备',
        '家用机零部件', '刺绣机零部件', '铺布设备零部件',
        # 产业用纺织品
        '产业用纺织品',
        # 纤维
        '玄武岩纤维', '化学纤维', '棉纤维', '聚丙烯腈纤维',
        # 纺织机械
        '针织横机', '环锭捻线机', '纺织机械',
        # 非织造布
        '非织造布',
        # 碳化装备
        '纤维碳化', '纤维碳化生产成套装备',
        # 服装用人体测量
        '服装用人体测量',
        # 棉纺织产品
        '棉纺织', '纺织纤维', '原棉', '棉纱', '棉布', '棉织物',
        '环锭纺纱', '紧密纺纱', '转杯纺纱', '喷气纺纱', '涡流纺纱',
        '精梳', '粗梳', '并条', '粗纱', '细纱', '络筒',
        '本色纱', '色纺纱', '染色纱', '漂白纱',
        '本色布', '色织布', '染色布', '漂白布', '印花布',
        # 化纤
        '再生纤维素纤维', '合成纤维', '人造纤维',
    ]

    # 从 PDF 文件名推断
    filenames = [
        '丝绸术语', '缝制机械术语', '产业用纺织品分类',
        '玄武岩纤维分类分级及代号', '化学纤维产品分类',
        '棉纤维术语分类和编码', '针织横机术语',
        '聚丙烯腈纤维生产成套装备术语',
        '棉纺织产品术语', '纺织品化学纤维产品术语',
        '环锭捻线机', '非织造布工厂设计规范',
        '服装用人体测量基准点',
    ]

    print(f"   标准术语: 约 {len(standard_domains)} 个（从标准目录编录）")

    return standard_domains


# ============================================================
# 合并 + 去重 + 分类
# ============================================================

def load_existing_lexicon():
    """加载所有已存在于 lexicon_v2 中的词条。"""
    if not LEXICON_PATH.exists():
        return set()

    with open(LEXICON_PATH, encoding='utf-8') as f:
        lex = yaml.safe_load(f)

    existing = set()

    def collect(val):
        if isinstance(val, str) and len(val) >= 2:
            existing.add(val)
        elif isinstance(val, list):
            for item in val:
                collect(item)
        elif isinstance(val, dict):
            for k, v in val.items():
                if k not in ('description', 'source', 'note',
                             '产业定位', '政策类型', 'policy_count', 'cluster_info'):
                    collect(k)
                    collect(v)

    for layer_key, layer_data in lex.get('layers', {}).items():
        collect(layer_data.get('terms', {}))

    return existing


def classify_term(term):
    """启发式分类: 将术语分配到 Layer 3 的八段分类。

    返回 (category_key, subcategory) 元组。
    """
    # 1. 原料端
    raw_material_keywords = ['纤维', '丝', '棉', '麻', '毛', '绒', '茧', '蚕', '羊毛',
                              '羊绒', '驼毛', '牦牛', '竹纤维', '木棉', '黄麻', '汉麻',
                              '涤纶', '锦纶', '氨纶', '腈纶', '丙纶', '维纶', '氯纶',
                              '莱赛尔', '莫代尔', '粘胶', '醋酸', '铜氨',
                              '纱', '线', '丝束', '长丝', '短纤', '短纤维',
                              '化学纤维', '合成纤维', '再生纤维', '人造纤维',
                              '碳纤维', '芳纶', '玄武岩', '聚酯', '聚酰胺', '聚丙烯腈',
                              '回收纤维', '再生涤纶', '原液着色', '色母粒']
    if any(kw in term for kw in raw_material_keywords):
        if any(kw in term for kw in ['合成', '涤纶', '锦纶', '氨纶', '腈纶', '丙纶', '维纶', '氯纶', '聚酯', '聚酰胺', '聚丙烯腈']):
            return ('layer_3_textile_chain', '1_原料端 / 合成纤维')
        if any(kw in term for kw in ['再生', '莱赛尔', '莫代尔', '粘胶', '醋酸', '铜氨', '竹浆', '天丝']):
            return ('layer_3_textile_chain', '1_原料端 / 再生纤维素纤维')
        if any(kw in term for kw in ['碳纤维', '芳纶', '玄武岩', '超高分子量']):
            return ('layer_3_textile_chain', '1_原料端 / 高性能纤维')
        return ('layer_3_textile_chain', '1_原料端 / 天然纤维')

    # 2. 纺纱
    spinning_keywords = ['纺纱', '纺', '并条', '粗纱', '细纱', '络筒', '捻线',
                          '环锭', '紧密纺', '转杯纺', '喷气纺', '涡流纺', '气流纺',
                          '精梳', '粗梳', '梳理', '并卷', '条卷',
                          '纱线', '单纱', '股线', '筒子纱', '绞纱']
    if any(kw in term for kw in spinning_keywords):
        return ('layer_3_textile_chain', '2_纺纱工艺')

    # 3. 织造
    weaving_keywords = ['织造', '织', '针织', '梭织', '经编', '纬编', '机织',
                         '横机', '圆机', '织机', '提花', '绣花',
                         '面料', '织物', '坯布', '色织', '牛仔']
    if any(kw in term for kw in weaving_keywords):
        return ('layer_3_textile_chain', '3_织造工艺')

    # 4. 非织造
    nonwoven_keywords = ['非织造', '无纺', '针刺', '水刺', '热轧', '纺粘', '熔喷',
                          '非织造布', '无纺布']
    if any(kw in term for kw in nonwoven_keywords):
        return ('layer_3_textile_chain', '4_非织造')

    # 5. 染整
    dyeing_keywords = ['染', '印花', '整理', '染色', '漂白', '丝光', '预缩',
                        '柔软整理', '抗皱', '防缩', '阻燃', '防水', '防油', '防污',
                        '扎染', '蜡染', '植物染料', '涂料', '后整理']
    if any(kw in term for kw in dyeing_keywords):
        return ('layer_3_textile_chain', '5_染整')

    # 6. 终端品类 / 服装
    product_keywords = ['服装', '服饰', '内衣', '衬衫', '裤', '裙', '上衣', '外套',
                         '西服', '夹克', '大衣', '羽绒服', 'T恤', 'POLO',
                         '袜', '帽', '手套', '围巾', '领带', '披肩',
                         '家纺', '床', '被', '枕', '毯', '毛巾', '浴',
                         '窗帘', '桌布', '地毯', '铺地',
                         '家居服', '睡衣', '运动服', '校服', '工装', '职业装',
                         '丝绸服装', '丝针织', '真丝', '新娘',
                         '包', '袋', '箱',
                         '产业用纺织品', '土工', '过滤', '汽车用', '医用']
    if any(kw in term for kw in product_keywords):
        return ('layer_3_textile_chain', '6_终端品类 / 流通渠道')

    # 7. 设备
    equipment_keywords = ['机', '设备', '装置', '仪器', '检测', '试验',
                           '缝纫机', '绣花机', '横机', '络筒机', '细纱机',
                           '梳理机', '并条机', '粗纱机', '捻线机', '整经机',
                           '浆纱机', '织机', '定型机', '烘干机', '拉幅机']
    if any(kw in term for kw in equipment_keywords):
        return ('layer_3_textile_chain', '7_设备与检测仪器')

    # 8. 智能制造 / 绿色 —— fallback, less specific
    smart_green = ['智能制造', '数字化', '绿色', '低碳', '零碳', '循环', '节能',
                   '环保', '碳排放', '清洁生产', '再生', '回收']
    if any(kw in term for kw in smart_green):
        return ('layer_3_textile_chain', '8_智能制造与绿色低碳')

    # Default: cross-domain
    return ('layer_5_cross_domain', '通用政策关键词（非排除）')


def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    parser = argparse.ArgumentParser(description="外部资料源批量术语提取")
    parser.add_argument("--output", type=str, default=str(OUTPUT_PATH),
                        help="输出 CSV 路径")
    args = parser.parse_args()

    print("=" * 60)
    print("📂 外部资料源批量术语提取")
    print("=" * 60)

    # 加载现有词典
    print("\n📚 加载现有词典...")
    existing = load_existing_lexicon()
    print(f"   已有词条: {len(existing)}")

    # === 数据源 1: HS Code ===
    print("\n📦 数据源 1: HS 商品编码表 (Ch50-63)...")
    hs_entries, hs_terms = extract_hs_terms()

    # === 数据源 2: 专精特新特色产品 ===
    print("\n📦 数据源 2: 前五批专精特新特色产品...")
    products, product_terms = extract_product_terms()

    # === 数据源 3: 行业分类 ===
    print("\n📦 数据源 3: 国民经济行业分类...")
    industry_terms = extract_industry_class_terms()

    # === 数据源 4: 标准术语 ===
    print("\n📦 数据源 4: GB/T 标准名称术语...")
    standard_terms = extract_standard_terms()

    # === 合并去重 ===
    print("\n" + "=" * 60)
    print("🔀 合并 + 去重 + 分类")
    print("=" * 60)

    all_candidates = set()
    sources = {}

    def add_terms(terms, source_label):
        for t in terms:
            t = t.strip()
            if len(t) >= 2 and re.match(r'^[一-鿿A-Za-z0-9·/\-]+$', t):
                # 跳过纯数字和常见噪声
                if re.match(r'^[\dA-Za-z]+$', t):
                    continue
                if len(t) == 2 and t[0] == t[1]:  # 叠字无意义
                    continue
                # 跳过常见虚词噪声
                if t in {'其他', '包括', '除外的', '不论是否', '未列名', '非供'}:
                    continue
                if t not in all_candidates:
                    all_candidates.add(t)
                    sources[t] = source_label

    add_terms(hs_terms, 'HS编码表')
    add_terms(product_terms, '专精特新产品')
    add_terms(industry_terms, '行业分类')
    add_terms(standard_terms, 'GB/T标准')

    print(f"   合并后候选: {len(all_candidates)} 个")

    # 去重（与现有词典）
    new_terms = sorted(all_candidates - existing)
    print(f"   去重后净增: {len(new_terms)} 个")

    # 分类
    print("\n📋 分类结果:")
    classified = {}
    for t in new_terms:
        layer, cat = classify_term(t)
        classified.setdefault((layer, cat), []).append(t)

    for (layer, cat), terms in sorted(classified.items()):
        print(f"   {layer}/{cat}: {len(terms)} terms")
        for t in sorted(terms)[:5]:
            print(f"     - {t}")
        if len(terms) > 5:
            print(f"     ... 等 {len(terms)} 个")

    # === 保存 ===
    output_path = Path(args.output)
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['term', 'length', 'source', 'layer', 'category', 'already_in_lexicon'])
        for t in sorted(new_terms):
            layer, cat = classify_term(t)
            writer.writerow([t, len(t), sources.get(t, 'unknown'), layer, cat, 'False'])
        # Also write existing terms that appeared in sources (for reference)
        for t in sorted(all_candidates & existing):
            writer.writerow([t, len(t), sources.get(t, 'unknown'), '—', '已在词典中', 'True'])

    print(f"\n💾 保存: {output_path}")
    print(f"   净增候选: {len(new_terms)}")
    print(f"   已在词典: {len(all_candidates & existing)}")
    print(f"   总计: {len(all_candidates)}")

    print()
    print("下一步:")
    print("  1. 人工审阅 source_enrichment.csv，逐条标记 manual_verdict (✅/❌)")
    print("  2. 运行 merge 脚本将通过的词条合并入 lexicon_v2.yaml")
    print("  3. 重新导出 jieba_dict.txt")


if __name__ == "__main__":
    main()
