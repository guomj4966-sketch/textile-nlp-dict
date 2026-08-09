"""从专精特新企业案例资料中提取纺织产业链术语，扩充第三层词库。

运行方式:
    python scripts/nlp_dict/extract_from_zhuanjingtexin.py
    python scripts/nlp_dict/extract_from_zhuanjingtexin.py --output data/layer_3_extended.yaml

依赖:
    docx, openpyxl, pdfplumber, jieba, pyyaml
    pip install python-docx openpyxl pdfplumber jieba pyyaml

输出:
    - scripts/nlp_dict/data/layer_3_textile_chain_extended.yaml

数据源:
    - OneDrive/专精特新平台/（外部数据，只读，通过相对路径定位）
"""

import sys
import re
import math
import argparse
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import jieba
import yaml

jieba.setLogLevel(20)

# ============================================================
# 配置
# ============================================================

# 从脚本位置向上4级到 OneDrive 根目录，再进入"专精特新平台"
_ONEDRIVE = Path(__file__).parent.parent.parent.parent
ZJT_DIR = _ONEDRIVE / "专精特新平台"
DATA_DIR = Path(__file__).parent / "data"
EXISTING_LEXICON_PATH = DATA_DIR / "lexicon_v1.yaml"

# jieba 加载已有词典
if (DATA_DIR / "jieba_dict.txt").exists():
    jieba.load_userdict(str(DATA_DIR / "jieba_dict.txt"))


# ============================================================
# 产业链术语分类框架（八段式）
# ============================================================

CHAIN_FRAMEWORK = {
    "1_原料端": {
        "description": "天然纤维/化学纤维/再生纤维/功能性纤维",
        "subcategories": {
            "天然纤维": ["棉花", "亚麻", "苎麻", "桑蚕丝", "柞蚕丝", "山羊绒", "绵羊毛", "羊驼毛", "马海毛", "牦牛绒",
                       "竹纤维", "木棉", "黄麻", "汉麻", "罗布麻", "棕榈纤维", "椰壳纤维", "菠萝纤维", "莲纤维"],
            "再生纤维素纤维": ["莱赛尔", "莫代尔", "粘胶纤维", "铜氨纤维", "醋酸纤维", "竹浆纤维", "天丝",
                           "纽代尔", "丽赛", "赛得利", "优可丝", "雅赛尔", "丝丽雅",
                           "离子液体法再生纤维素", "NMMO法", "新型溶剂法"],
            "合成纤维": ["涤纶", "锦纶", "氨纶", "腈纶", "丙纶", "维纶", "氯纶",
                       "聚酯纤维", "聚酰胺纤维", "聚丙烯腈纤维", "聚氨酯弹性纤维",
                       "PTT", "PBT", "PPS", "PE", "PP", "PA6", "PA66",
                       "超细旦", "细旦", "异形截面", "中空纤维", "复合纤维", "双组分纤维",
                       "原液着色", "熔体直纺", "切片纺"],
            "生物基与可降解": ["聚乳酸", "PLA", "生物基聚酰胺", "生物基涤纶", "PHA", "PHBV",
                           "壳聚糖纤维", "海藻纤维", "大豆蛋白纤维", "牛奶蛋白纤维",
                           "聚羟基脂肪酸酯", "生物基PTT", "生物基PE"],
            "高性能纤维": ["碳纤维", "芳纶", "超高分子量聚乙烯", "玄武岩纤维", "玻璃纤维",
                        "聚酰亚胺", "聚苯硫醚", "聚四氟乙烯", "芳砜纶", "陶瓷纤维",
                        "碳化硅纤维", "氧化铝纤维", "硼纤维", "PBO纤维", "M5纤维"],
            "功能性/差别化": ["抗菌纤维", "阻燃纤维", "导电纤维", "防紫外线纤维", "远红外纤维",
                          "负离子纤维", "吸湿排汗", "凉感纤维", "蓄热纤维", "相变纤维",
                          "光致变色", "温致变色", "夜光纤维", "磁性纤维", "消臭纤维",
                          "石墨烯纤维", "纳米纤维", "静电纺丝"],
            "再生/循环纤维": ["再生涤纶", "再生锦纶", "再生棉", "再生毛", "再生氨纶",
                           "废旧纺织品回收", "物理法再生", "化学法再生", "瓶片纺丝",
                           "消费后回收", "GRS认证"],
        },
    },
    "2_纺纱端": {
        "description": "纺纱工艺/纱线品种/纱线结构",
        "subcategories": {
            "纺纱工艺": ["环锭纺", "紧密纺", "赛络纺", "涡流纺", "气流纺", "转杯纺", "喷气纺",
                       "包芯纺", "包覆纺", "花式纺", "段彩纺", "竹节纺", "AB纺",
                       "嵌入式复合纺", "低扭矩纺", "柔洁纺", "聚纤纺", "扭妥纺"],
            "纱线结构": ["单纱", "股线", "包芯纱", "包覆纱", "花式纱", "段彩纱", "竹节纱",
                       "弹力纱", "高支纱", "低支纱", "精梳纱", "普梳纱", "半精纺",
                       "紧密纱", "涡流纱", "气流纱", "环锭纱",
                       "赛络紧密纺纱", "赛络菲尔纱", "长丝短纤复合纱"],
            "纱线品类": ["纯棉纱", "涤棉纱", "混纺纱", "色纺纱", "原液着色纱",
                       "氨纶包芯纱", "锦纶包覆纱", "涤纶缝纫线",
                       "全棉OE纱", "全棉环纺纱", "有机棉纱",
                       "麻灰纱", "雪花纱", "彩点纱", "AB竹节纱"],
            "前纺与准备": ["清梳联", "梳棉", "并条", "精梳", "粗纱", "细纱", "络筒",
                        "倍捻", "并线", "摇纱", "绞纱", "筒染", "绞染"],
        },
    },
    "3_织造端": {
        "description": "织造方式/面料结构/针织经编",
        "subcategories": {
            "梭织": ["平纹", "斜纹", "缎纹", "提花", "色织", "牛仔", "灯芯绒", "泡泡纱",
                   "双层组织", "三层组织", "表里换层", "起绒组织", "蜂巢组织",
                   "高支高密", "府绸", "牛津纺", "青年布", "华达呢", "卡其"],
            "针织": ["纬编", "经编", "单面", "双面", "罗纹", "双罗纹", "珠地",
                   "提花针织", "横机", "圆机", "无缝针织", "全成型",
                   "毛圈", "法式毛圈", "空气层", "三明治", "拉绒"],
            "非织造": ["水刺", "针刺", "热轧", "纺粘", "熔喷", "SMS", "SMMS",
                     "化学粘合", "热风", "湿法", "静电纺",
                     "土工布", "过滤材料", "擦拭布", "医用无纺布"],
            "编织与其他": ["编织", "钩编", "簇绒", "植绒", "蕾丝", "网眼",
                        "窄幅织带", "松紧带", "绳子", "缆绳"],
        },
    },
    "4_染整端": {
        "description": "染色/印花/整理/助剂",
        "subcategories": {
            "前处理": ["退浆", "煮练", "漂白", "丝光", "碱减量", "酶处理",
                     "生物酶前处理", "冷堆", "连续前处理", "短流程"],
            "染色": ["浸染", "轧染", "气流染色", "溢流染色", "经轴染色",
                   "球经染色", "绳状染色", "喷射染色", "高温高压染色",
                   "活性染料", "分散染料", "酸性染料", "阳离子染料",
                   "低温染色", "无水染色", "超临界CO2染色", "溶剂染色",
                   "原位矿化染色", "汽介数字染色", "植物染料染色",
                   "天然染料", "生物质染料", "微生物色素"],
            "印花": ["数码印花", "丝网印花", "滚筒印花", "转移印花", "热转印",
                   "喷墨印花", "涂料印花", "烂花", "拔染", "防染",
                   "发泡印花", "植绒印花", "烫金", "烫银",
                   "功能性印花", "渐变印花", "3D印花"],
            "后整理": ["免烫", "防缩", "防皱", "丝光", "柔软", "硬挺",
                     "阻燃整理", "防水整理", "防油整理", "防污整理", "易去污",
                     "吸湿排汗", "速干", "凉感", "蓄热", "保暖",
                     "远红外", "防紫外线", "抗菌整理", "防螨", "防蚊",
                     "消臭", "芳香", "芦荟", "维生素", "胶原蛋白",
                     "负离子", "磁疗", "远红外",
                     "抗静电", "防辐射", "电磁屏蔽",
                     "涂层", "覆膜", "贴合", "层压",
                     "纳米整理", "等离子体", "超声波"],
            "印染助剂": ["前处理剂", "染色助剂", "印花助剂", "后整理剂",
                       "匀染剂", "固色剂", "渗透剂", "消泡剂",
                       "生物基助剂", "绿色助剂", "环保助剂",
                       "柔软剂", "防水剂", "阻燃剂", "抗菌剂"],
        },
    },
    "5_成衣/缝制端": {
        "description": "服装制造/缝制/裁剪/后道",
        "subcategories": {
            "裁剪": ["自动裁床", "激光裁剪", "超声波裁剪", "智能排版", "CAD排版",
                   "单层裁剪", "多层裁剪", "对格对条"],
            "缝制": ["平缝", "包缝", "绷缝", "链缝", "锁眼", "钉扣",
                   "模板缝制", "自动缝制", "机器人缝制",
                   "无缝", "热封", "超声波粘合", "高频焊接",
                   "吊挂系统", "柔性生产", "单件流", "模块化生产"],
            "后道": ["整烫", "检针", "包装", "折叠", "挂装",
                   "水洗", "酵素洗", "石磨", "砂洗", "免烫处理"],
            "辅料": ["腰衬", "粘合衬", "纽扣", "拉链", "魔术贴", "织标",
                   "吊牌", "包装袋", "衣架", "衬布", "垫肩"],
        },
    },
    "6_终端品类": {
        "description": "服装/家纺/产业用纺织品细分品类",
        "subcategories": {
            "服装品类": ["男装", "女装", "童装", "婴装", "运动服", "户外服",
                       "冲锋衣", "羽绒服", "大衣", "西服", "衬衫",
                       "T恤", "卫衣", "针织衫", "毛衣", "连衣裙",
                       "工装", "职业装", "校服", "防护服", "军服",
                       "内衣", "文胸", "内裤", "袜品", "家居服",
                       "泳装", "瑜伽服", "骑行服", "滑雪服"],
            "家纺品类": ["床品套件", "被芯", "枕芯", "床垫", "毛巾", "浴巾",
                       "窗帘", "沙发套", "桌布", "地毯", "毛毯",
                       "蚊帐", "凉席", "抱枕", "靠垫",
                       "功能性家纺", "智能家纺", "抗菌家纺"],
            "产业用纺织品": ["土工布", "过滤材料", "车用纺织品", "汽车内饰",
                           "医用纺织品", "口罩", "防护服", "绷带",
                           "卫生用品", "纸尿裤", "湿巾",
                           "航天", "军工", "特种", "绳索", "渔网",
                           "篷盖", "遮阳", "帐篷", "帆布",
                           "建筑", "管道", "风电", "船用",
                           "输送带", "传动带", "密封材料"],
            "功能性终端": ["防晒", "防雨", "防风", "透气", "速干",
                        "抗菌", "防臭", "防螨", "防蚊",
                        "智能温控", "智能心率", "柔性电子",
                        "自清洁", "形状记忆", "可穿戴"],
        },
    },
    "7_智能制造": {
        "description": "数字化/自动化/AI与纺织制造融合",
        "subcategories": {
            "数字系统": ["ERP", "MES", "PLM", "SCM", "WMS", "APS",
                       "工业互联网平台", "数字孪生", "云计算", "大数据",
                       "雅典娜纺织工业互联网平台"],
            "智能装备": ["自动落纱", "自动络筒", "自动搬运AGV", "巡检机器人",
                       "智能验布", "机器视觉检测", "AI质检",
                       "数码印花机", "Single-Pass数码印花",
                       "全自动绣花机", "智能吊挂"],
            "数据与AI": ["大模型", "机器学习", "深度学习", "图像识别",
                       "预测性维护", "能耗管理", "碳数据管理",
                       "全生命周期管理", "数字护照DPP"],
        },
    },
    "8_绿色低碳循环": {
        "description": "节能/减排/零碳/循环/认证",
        "subcategories": {
            "减排技术": ["光伏", "分布式光伏", "屋顶光伏", "余热回收",
                       "中水回用", "废水处理", "废气治理", "VOCs治理",
                       "污泥处理", "固废资源化", "危废处置",
                       "生物法处理", "深井曝气", "膜处理"],
            "绿色认证": ["GRS", "GOTS", "OCS", "BCI", "OEKO-TEX",
                       "ZDHC", "Higg Index", "碳足迹", "碳标签",
                       "零碳工厂", "绿色工厂", "绿色设计产品",
                       "EPD", "PEF", "LCA"],
            "循环模式": ["循环经济", "闭环回收", "纺织到纺织", "物理回收",
                       "化学回收", "废旧纺织品再生", "二手交易",
                       "租赁模式", "共享制造", "产品即服务"],
        },
    },
}

# 框架中的所有术语集合（用于快速查找）
FRAMEWORK_TERMS: set = set()


def _collect_framework_terms():
    global FRAMEWORK_TERMS
    if FRAMEWORK_TERMS:
        return FRAMEWORK_TERMS

    def collect(d):
        for k, v in d.items():
            if isinstance(v, dict):
                if "subcategories" in v:
                    for subcat, terms in v["subcategories"].items():
                        if isinstance(terms, list):
                            FRAMEWORK_TERMS.update(t.strip() for t in terms)
                else:
                    collect(v)
            elif isinstance(v, list):
                FRAMEWORK_TERMS.update(t.strip() for t in v)

    collect(CHAIN_FRAMEWORK)
    return FRAMEWORK_TERMS


_collect_framework_terms()


# ============================================================
# 数据源读取
# ============================================================


def read_excel_products(path):
    """从产品技术目录汇总表读取所有产品名。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ⚠️ openpyxl 未安装，跳过 Excel 文件")
        return []

    wb = load_workbook(path, data_only=True)
    all_products = []

    # 汇总 sheet
    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[1]:
                all_products.append(str(row[1]).strip())

    # 2024-2026 各sheet
    for sn in ["2024", "2025", "2026"]:
        if sn in wb.sheetnames:
            ws = wb[sn]
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if row and any("编号" in str(c) for c in row if c):
                    header_row = row
                    break
            if header_row:
                data_start = 2
                for row in ws.iter_rows(min_row=data_start, values_only=True):
                    if row and len(row) >= 3:
                        name_col = 2 if len(row) > 2 else 1
                        prod_name = str(row[name_col]).strip() if row[name_col] else ""
                        if prod_name and len(prod_name) > 3:
                            all_products.append(prod_name)

    wb.close()
    return all_products


def read_excel_characteristic_products(path):
    """从特色产品表读取产品名和描述。"""
    try:
        import pandas as pd
    except ImportError:
        print("  ⚠️ pandas 未安装，跳过 特色产品表")
        return [], []

    xl = pd.ExcelFile(path)
    products = []
    descriptions = []

    if "前5批特色产品" in xl.sheet_names:
        df = pd.read_excel(path, sheet_name="前5批特色产品")
        if "特色产品名称" in df.columns:
            for v in df["特色产品名称"].dropna():
                products.append(str(v).strip().replace("_x000D_", ""))
        if "特色产品介绍" in df.columns:
            for v in df["特色产品介绍"].dropna():
                descriptions.append(str(v).strip().replace("_x000D_", ""))

    return products, descriptions


def read_docx_files(directory, max_files=None):
    """从文件夹读取所有 docx 全文。"""
    try:
        from docx import Document
    except ImportError:
        print("  ⚠️ python-docx 未安装，跳过 docx 文件")
        return []

    texts = []
    files = list(Path(directory).glob("*.docx")) + list(Path(directory).glob("*.doc"))
    if max_files:
        files = files[:max_files]

    for fp in files:
        try:
            doc = Document(str(fp))
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            if paragraphs:
                texts.append("\n".join(paragraphs))
        except Exception:
            pass  # 跳过无法读取的文件

    return texts


def read_pdf_files(directory):
    """从文件夹读取所有 PDF 全文（跳过纯扫描版）。"""
    try:
        import pdfplumber
    except ImportError:
        print("  ⚠️ pdfplumber 未安装，跳过 PDF 文件")
        return []

    texts = []
    for fp in Path(directory).glob("*.pdf"):
        total_chars = 0
        try:
            with pdfplumber.open(str(fp)) as pdf:
                for page in pdf.pages:
                    txt = page.extract_text() or ""
                    total_chars += len(txt)
                if total_chars > 500:  # 跳过扫描版（无文字或极少文字）
                    full = []
                    for page in pdf.pages:
                        t = page.extract_text() or ""
                        if t.strip():
                            full.append(t)
                    texts.append("\n".join(full))
                    print(f"  ✅ {fp.name}: {total_chars:,} 字 ({len(pdf.pages)} 页)")
                else:
                    print(f"  ⏭ {fp.name}: 仅为 {total_chars} 字（扫描版，跳过）")
        except Exception as e:
            print(f"  ⚠️ {fp.name}: 读取失败 ({e})")

    return texts


# ============================================================
# 术语发现：在文本中找框架外的新术语
# ============================================================


def discover_new_terms(texts, framework_terms, min_len=2, min_freq=3):
    """在文本中发现不在框架中的新术语。

    策略：用 jieba 分词，过滤出不在框架中的高频中文词，
    再通过启发式规则筛选可能有领域价值的候选。
    """
    term_counter = Counter()

    for text in texts:
        words = jieba.lcut(text)
        chinese_words = [
            w.strip() for w in words
            if len(w.strip()) >= min_len
            and re.match(r'^[一-鿿a-zA-Z0-9/+]+$', w.strip())
            and w.strip() not in framework_terms
        ]
        term_counter.update(chinese_words)

    # 过滤
    new_terms = {}
    for term, freq in term_counter.most_common():
        if freq < min_freq:
            break

        # 排除纯数字
        if re.match(r'^[\d\.\+\-/]+$', term):
            continue

        # 排除单字
        if len(term) < 2:
            continue

        # 排除纯停用词
        if term in STOP_WORDS:
            continue

        # 排除通用动词/形容词
        if _is_generic(term):
            continue

        new_terms[term] = freq

    return new_terms


# 停用词
STOP_WORDS = {
    '根据', '按照', '关于', '有关', '相关', '应当', '可以', '予以',
    '负责', '确保', '组织', '协调', '建立', '制定', '实施', '执行',
    '落实', '推进', '开展', '加强', '完善', '进一步', '持续', '不断',
    '着力', '加大', '提高', '降低', '减少', '增加', '扩大',
    '目前', '已经', '正在', '计划', '准备', '预计', '预计到',
    '通过', '采用', '利用', '应用', '使用', '包括', '包含',
    '主要', '其中', '其中以', '同时', '此外', '另外',
    '以上', '以下', '上述', '下列', '如下',
    '为', '的是', '以及', '及其', '并按', '根据各',
    '发展', '建设', '提升', '重点', '领域', '工作', '支持',
    '服务', '产业', '开展', '国家', '企业', '优化',
    '市场', '创新', '实现', '形成', '推动',
    '这些', '它们', '所有', '整个', '全部', '部分',
    '第一', '第二', '第三', '事项', '内容', '情况',
    '产品', '技术', '行业', '项目', '单位', '部门',
    '该产品', '该项目', '采用', '生产', '加工',
    '等方面', '等领域', '等行业', '的研发', '的研制',
}


def _is_generic(term):
    """判断是否为过于通用的词汇。"""
    generic_suffixes = [
        '能力', '水平', '效率', '质量', '效益', '效果',
        '程度', '速度', '力度', '强度', '广度', '深度',
        '问题', '需求', '目标', '任务', '措施', '方法',
        '方案', '规划', '政策', '法规', '标准', '规范',
        '工作', '业务', '流程', '过程', '环节', '步骤',
        '设备', '装置', '系统', '平台', '工具', '软件',
        '材料', '原料', '资源', '能源', '资金', '资产',
    ]
    for suf in generic_suffixes:
        if term.endswith(suf):
            return True
    return False


# ============================================================
# 提取脚本主体
# ============================================================


def main():
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    parser = argparse.ArgumentParser(description="从专精特新案例提取产业链术语")
    parser.add_argument("--output", default="data/layer_3_textile_chain_extended.yaml",
                        help="输出文件路径")
    parser.add_argument("--max-files", type=int, default=0,
                        help="最多读取 N 个 docx 文件（0=全部）")
    args = parser.parse_args()

    print("=" * 60)
    print("🧵 从专精特新案例提取纺织产业链术语")
    print("=" * 60)

    framework_terms = _collect_framework_terms()
    print(f"\n📚 框架已有术语: {len(framework_terms)} 个")

    all_texts = []
    all_product_names = []

    # ---- 1. Excel 产品名 ----
    print("\n📊 [1/4] 读取 Excel 产品名...")
    products_xlsx = ZJT_DIR / "新产品（文字）" / "2024-2026产品技术目录汇总.xlsx"
    if products_xlsx.exists():
        prods = read_excel_products(products_xlsx)
        all_product_names.extend(prods)
        print(f"   ✅ {len(prods)} 个产品名")
    else:
        print(f"   ⚠️ 文件不存在: {products_xlsx}")

    # ---- 2. 特色产品表 ----
    print("\n📊 [2/4] 读取特色产品表...")
    char_prod_path = ZJT_DIR / "前五批纺织行业专精特新企业汇总（培育入库）.xlsx"
    if char_prod_path.exists():
        cp_products, cp_descs = read_excel_characteristic_products(char_prod_path)
        all_product_names.extend(cp_products)
        all_texts.extend(cp_descs)
        total_chars = sum(len(d) for d in cp_descs)
        print(f"   ✅ {len(cp_products)} 个特色产品名 + {len(cp_descs)} 条产品介绍 ({total_chars:,}字)")
    else:
        print(f"   ⚠️ 文件不存在: {char_prod_path}")

    # ---- 3. 案例 docx ----
    print("\n📄 [3/4] 读取案例 docx...")
    case_dirs = [
        ZJT_DIR / "企业案例" / "企业案例-2022发展报告",
        ZJT_DIR / "企业案例" / "企业案例-2026绿色发展",
        ZJT_DIR / "新产品（文字）" / "2024汇编材料文字版",
        ZJT_DIR / "新产品（文字）" / "2025汇编材料文字版",
        ZJT_DIR / "新产品（文字）" / "2026绿色产品技术",
    ]
    docx_total = 0
    for d in case_dirs:
        if d.exists():
            docs = read_docx_files(str(d), max_files=args.max_files if args.max_files else None)
            all_texts.extend(docs)
            total_chars = sum(len(doc) for doc in docs)
            print(f"   ✅ {d.name}: {len(docs)} 个文件 ({total_chars:,}字)")
            docx_total += len(docs)
        else:
            print(f"   ⚠️ 目录不存在: {d}")
    print(f"   合计 docx: {docx_total} 个, 总文本: {sum(len(t) for t in all_texts):,}字")

    # ---- 4. PDF 报告 ----
    print("\n📑 [4/4] 读取研究报告 PDF...")
    pdf_dir = ZJT_DIR / "研究报告"
    if pdf_dir.exists():
        pdf_texts = read_pdf_files(str(pdf_dir))
        for pt in pdf_texts:
            all_texts.append(pt)
    else:
        print(f"   ⚠️ 目录不存在: {pdf_dir}")

    print(f"\n📦 总计: {len(all_product_names)} 个产品名 + {len(all_texts)} 个文本段落")

    # ---- 5. 从产品名中提取术语 ----
    print("\n" + "=" * 60)
    print("🔍 术语发现...")
    print("=" * 60)

    # 5a: 从产品名中提取框架外的新术语
    print(f"\n📋 从 {len(all_product_names)} 个产品名中提取术语...")
    discovered_from_products = set()
    for pn in all_product_names:
        # 产品名通常是最浓缩的术语来源
        # 移除版本号和编号
        pn_clean = re.sub(r'[\d\.\-\+\（\）\(\)]+', ' ', pn)
        pn_clean = re.sub(r'\s+', ' ', pn_clean).strip()

    # 5b: 从全文文本中分词发现
    print(f"📋 从 {len(all_texts)} 个文本中分词发现...")

    total_chars = sum(len(t) for t in all_texts)
    print(f"   总文本: {total_chars:,} 字")

    discovered = discover_new_terms(all_texts, framework_terms, min_len=2, min_freq=3)
    print(f"   发现新候选术语: {len(discovered)} 个")

    # 5c: 展示 Top 新词
    print(f"\n📊 新发现候选术语 Top 50:")
    count = 0
    for term, freq in discovered.items():
        # 只展示在框架外的，且与纺织相关的
        if has_textile_relevance(term):
            print(f"   {term:<20} (频次: {freq})")
            count += 1
            if count >= 50:
                break

    if count == 0:
        print("   (需要人工审阅，见输出文件)")

    # ---- 6. 整理输出 ----
    print("\n" + "=" * 60)
    print("💾 保存结果...")
    print("=" * 60)

    # 构建扩展版第三层词典
    extended = {
        "meta": {
            "version": "2.0-extended",
            "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "source": "专精特新平台/企业案例 + 新产品产品名 + 研究报告",
            "description": "纺织产业链术语体系扩展版——从300+专精特新企业案例和新产品资料中提取",
            "framework_base_terms": len(framework_terms),
            "new_discovered_count": len(discovered),
        },
        "chain_framework": CHAIN_FRAMEWORK,
        "discovered_terms": {
            term: freq for term, freq in sorted(
                discovered.items(), key=lambda x: -x[1]
            )
        },
    }

    output_path = DATA_DIR / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        yaml.dump(
            extended, f,
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
            width=120,
        )

    print(f"💾 {output_path}")
    print(f"   八段框架术语: {len(framework_terms)} 个")
    print(f"   新发现候选: {len(discovered)} 个")

    # 统计分布
    chain_count = 0
    for section_name, section in CHAIN_FRAMEWORK.items():
        sub_total = 0
        for subcat_name, terms in section.get("subcategories", {}).items():
            sub_total += len(terms)
        chain_count += sub_total
        print(f"   {section_name:<20} {sub_total:>4} 词 — {section['description']}")

    print(f"\n   框架术语合计: {chain_count}")
    print(f"   新发现候选:   {len(discovered)}")
    print(f"   总计入库:     {chain_count + len(discovered)}")

    print("\n下一步:")
    print("  1. 人工审阅 discovered_terms 中的候选词")
    print("  2. 将确认通过的候选词归入 chain_framework 对应子类")
    print("  3. 运行 merge_lexicon.py --category textile_chain 更新词典")
    print("  4. 更新 jieba_dict.txt")

    return extended


def has_textile_relevance(term):
    """粗略判断一个候选术语是否与纺织相关。"""
    # 纺织核心字
    textile_chars = '纺织服装棉麻丝毛纤维纱布染印经编缝纫缫浆氨涤锦腈纶维碳芳'
    if any(c in term for c in textile_chars):
        return True

    # 纺织关键词
    textile_keywords = [
        '面', '料', '色', '纱', '线', '绳', '带', '絮', '绒',
        '衫', '裤', '裙', '巾', '毯', '帘', '垫', '枕', '被',
        '袖', '领', '扣', '链', '衬',
    ]
    if any(kw in term for kw in textile_keywords):
        return True

    return False


if __name__ == "__main__":
    main()
